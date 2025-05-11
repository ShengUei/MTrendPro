import re
import yfinance as yf  # Using Yahoo Finance API
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers
import configparser
import os
import sys

def load_config(config_file='config.ini'):
    """
    Load configuration from the specified ini file
    """
    if not os.path.exists(config_file):
        print(f"Error: Configuration file '{config_file}' not found.")
        print(f"Please ensure '{config_file}' exists in the same directory as the script.")
        sys.exit(1)
        
    config = configparser.ConfigParser()
    config.read(config_file)
    
    # Validate required configuration
    required_sections = ['Excel']
    required_options = {'Excel': ['file_path', 'sheet_name']}
    
    for section in required_sections:
        if section not in config:
            print(f"Error: Required section '{section}' not found in the configuration file.")
            sys.exit(1)
        
        for option in required_options.get(section, []):
            if option not in config[section]:
                print(f"Error: Required option '{option}' not found in section '{section}'.")
                sys.exit(1)
    
    return config

# Load configuration
config = load_config()
excel_file_path = config['Excel']['file_path']
sheet_name = config['Excel']['sheet_name']
verbose = config.getboolean('Output', 'verbose', fallback=True)

# Read stock symbols from the specified sheet
df = pd.read_excel(excel_file_path, sheet_name=sheet_name)

# Assume stock codes are in the "Symbol" column
stock_symbols = df["Symbol"].tolist()

# Create dictionaries to store results
results = {}
market_caps = {}

# Get today's date as part of the column name
today = datetime.today().strftime('%Y-%m-%d')
column_name_close = f"{today}_Close"
column_name_market_cap = f"{today}_MarketCap(T)"  # MarketCap(T) indicates market capitalization in trillions

# Define a function to convert market capitalization values (handling B and T abbreviations, converting to T)
def parse_market_cap(market_cap):
    # Convert market cap from B (billions) to T (trillions)
    if market_cap is not None:
        return market_cap / 1000  # Convert B (billion dollars) to T (trillion dollars)
    return None

# Define a function to determine whether to fetch market cap based on asset type
def should_fetch_market_cap(ticker_info):
    """
    Determine whether to fetch market cap based on stock information
    """
    # If ticker_info is empty or has no quoteType, return False directly
    if not ticker_info or 'quoteType' not in ticker_info:
        return False
        
    # Determine based on quoteType
    quote_type = ticker_info.get('quoteType', '').lower()
    
    # ETFs and currencies don't need market cap, but cryptocurrencies do
    if quote_type in ['etf', 'mutualfund', 'currency']:
        return False
        
    # If the type is stock or cryptocurrency, fetch market cap
    if quote_type in ['equity', 'stock', 'cryptocurrency']:
        return True
        
    # Check for Taiwan ETFs with specific suffixes (like .TW)
    symbol = ticker_info.get('symbol', '')
    if '.TW' in symbol and ticker_info.get('fundFamily') is not None:
        return False
        
    # For uncertain types, check if marketCap field exists
    return 'marketCap' in ticker_info and ticker_info['marketCap'] is not None

# Loop through each stock symbol and fetch closing price and market cap
for stock_symbol in stock_symbols:
    try:
        # Use Yahoo Finance API to get stock information
        ticker = yf.Ticker(stock_symbol)
        
        # Get the latest market data
        ticker_info = ticker.info
        
        # Get closing price
        if 'regularMarketPrice' in ticker_info and ticker_info['regularMarketPrice'] is not None:
            close_price = ticker_info['regularMarketPrice']
            results[stock_symbol] = close_price
            if verbose:
                print(f"Successfully fetched {stock_symbol} price: {close_price}")
        else:
            print(f"Failed to fetch {stock_symbol} closing price")
        
        # Determine whether to fetch market capitalization
        if should_fetch_market_cap(ticker_info):
            # Get market cap (in USD)
            if 'marketCap' in ticker_info and ticker_info['marketCap'] is not None:
                # Yahoo Finance returns market cap as a raw number (without B/T), so we need to convert to T
                market_cap_in_billions = ticker_info['marketCap'] / 1000000000  # Convert to billions
                market_cap_in_trillions = market_cap_in_billions / 1000  # Convert to trillions
                market_caps[stock_symbol] = market_cap_in_trillions
                if verbose:
                    print(f"Successfully fetched {stock_symbol} market cap: {market_cap_in_billions}B → {market_cap_in_trillions}T")
            else:
                print(f"Failed to fetch {stock_symbol} market cap")
        else:
            asset_type = ticker_info.get('quoteType', 'Unknown Type')
            if verbose:
                print(f"Skipping market cap for {stock_symbol} (Asset Type: {asset_type})")
            
    except Exception as e:
        print(f"Error processing {stock_symbol}: {e}")

# Open Excel file
wb = load_workbook(excel_file_path)
ws = wb[sheet_name]

# Find the "Symbol" column and write date-tagged closing price and market cap columns next to it
symbol_col_idx = None
for col in ws.iter_cols(1, ws.max_column):
    if col[0].value == "Symbol":
        symbol_col_idx = col[0].column
        break

if symbol_col_idx is not None:
    # Write "2024-08-24_Close" and "2024-08-24_MarketCap(T)" columns next to the "Symbol" column
    close_col_idx = symbol_col_idx + 1
    market_cap_col_idx = symbol_col_idx + 2

    ws.cell(row=1, column=close_col_idx, value=column_name_close)  # Set closing price header
    ws.cell(row=1, column=market_cap_col_idx, value=column_name_market_cap)  # Set market cap header (in T units)

    for row in range(2, ws.max_row + 1):
        symbol = ws.cell(row=row, column=symbol_col_idx).value
        if symbol in results:
            ws.cell(row=row, column=close_col_idx, value=results[symbol])
            ws.cell(row=row, column=close_col_idx).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        if symbol in market_caps:
            ws.cell(row=row, column=market_cap_col_idx, value=market_caps[symbol])

# Save Excel file
wb.save(excel_file_path)

print("Data has been updated to the Excel file.")